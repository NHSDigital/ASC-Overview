from pathlib import Path
from typing import Callable, Union, cast
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from asc_overview.cms_tables import cms_filenames as name
from asc_overview.utilities.output_to_excel import write_dataframe_to_excel
from asc_overview.load_publications.overview_reader.overview_reader import (
    OverviewReader,
)


def output_all_cms_tables(overview_reader: OverviewReader, output_path: str):
    for (
        cms_filename,
        generator_function,
    ) in get_cms_generator_by_filename_dict(overview_reader).items():
        cms_dataframe = generator_function().pipe(format_pandas_object_for_cms_output)
        path_to_save_to = Path(output_path) / cms_filename
        output_cms_table(cms_dataframe, path_to_save_to)


def output_cms_table(cms_dataframe: pd.DataFrame, path_to_save_to: Path):
    cms_workbook = Workbook()
    cms_worksheet = cms_workbook.active

    write_dataframe_to_excel(cms_worksheet, cms_dataframe, "A1")
    format_index_as_bold(cms_worksheet, len(cms_dataframe))
    format_columnns_as_wrapped(cms_worksheet, len(cms_dataframe.columns))

    cms_workbook.save(str(path_to_save_to))


def format_pandas_object_for_cms_output(cms_dataframe: Union[pd.DataFrame, pd.Series]):
    return cms_dataframe.reset_index().T.reset_index().T.replace("index", np.nan)


def format_index_as_bold(cms_worksheet: Worksheet, length_of_index: int):
    for i in range(1, length_of_index + 1):
        cell_to_format = cast(Cell, cms_worksheet.cell(i, 1))
        cell_to_format.font = Font(bold=True)


def format_columnns_as_wrapped(cms_worksheet: Worksheet, length_of_columns: int):
    for i in range(1, length_of_columns + 1):
        cell_to_format = cast(Cell, cms_worksheet.cell(1, i))
        cell_to_format.alignment = Alignment(wrap_text=True)


def get_cms_generator_by_filename_dict(
    overview_reader: OverviewReader,
) -> dict[str, Callable]:
    return {
        name.NEW_REQUESTS_FOR_SUPPORT: overview_reader.get_new_requests_for_support,
        name.CHANGE_IN_NEW_REQUESTS: overview_reader.get_change_in_requests,
        name.NEW_REQUESTS_PER_POPULATION: overview_reader.get_new_requests_per_population,
        name.DOLS_REQUESTS: overview_reader.get_dols_requests,
        name.SAEFEGUARDING_REQUESTS: overview_reader.get_safeguarding_requests,
        name.CHANGE_IN_SAFEGUARDING_AND_DOLS: overview_reader.get_change_in_safeguarding_and_dols,
        name.SHORT_TERM_SUPPORT: overview_reader.get_short_term_support,
        name.WHAT_HAPPENED_NEXT: overview_reader.get_percent_what_happened_next,
        name.LONG_TERM_SUPPORT: overview_reader.get_long_term_support_per_population,
        name.LONG_TERM_SUPPORT_SETTING: overview_reader.get_long_term_support_setting,
        name.PRIMARY_SUPPORT_REASON: overview_reader.get_primary_support_reason,
        name.GROSS_CURRENT_EXPENDITURE: overview_reader.get_gross_current_expenditure,
        name.GROSS_CURRENT_EXPENDITURE_PER_CITIZEN: overview_reader.get_expenditure_per_citizen,
        name.LA_EXPENDITURE: overview_reader.get_proportion_la_expenditure,
        name.JOBS: overview_reader.get_jobs,
        name.VACANCIES: overview_reader.get_vacancies,
        name.STARTERS_AND_LEAVERS: overview_reader.get_starters_and_leavers,
        name.TYPE_OF_ROLE: overview_reader.get_job_role,
        name.SERVICE_USER_SATISFACTION: overview_reader.get_service_user_satisfaction,
        name.SERVICE_USER_THEMSELVES: overview_reader.get_service_user_feelings_themselves,
        name.SERVICE_USER_CHOICE: overview_reader.get_service_user_feelings_choice,
        name.CARER_SATISFACTION: overview_reader.get_carer_satisfaction,
        name.SERVICE_USER_QUALITY_OF_LIFE: overview_reader.get_service_user_quality_of_life,
        name.CARER_QUALITY_OF_LIFE: overview_reader.get_carer_quality_of_life,
        name.SAFEGUARDING_ENQUIRIES: overview_reader.get_safeguarding_enquiries,
        name.SERVICE_USER_SAFETY: overview_reader.get_service_user_feelings_safety,
        name.SERVICE_USER_SOCIAL_CONTACT: overview_reader.get_service_user_social_contact,
        name.CARER_SOCIAL_CONTACT: overview_reader.get_carer_social_contact,
    }
