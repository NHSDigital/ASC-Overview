from typing import Union
import numpy as np
import pandas as pd
from asc_overview.utilities.excel_config import ExcelWorkbookRange
from asc_overview.utilities.convert_cell_ref_to_numbers import (
    CellReferenceNumbers,
    convert_cell_ref_to_numbers_one_based,
    convert_cell_ref_to_numbers_zero_based,
)
from openpyxl.worksheet.worksheet import Worksheet


def load_workbook_range_as_dataframe(
    workbook_range: ExcelWorkbookRange, sheet: Worksheet
) -> pd.DataFrame:
    """
    This function will read in a range from an excel worksheet
    from the cell start to the cell end

    e.g cell_ref=C2, cell_end=F4 will read in data
    from the excel range C2:F4 with:
    - C2, C3 and C4 as the column names
    - C2, D2, E2 and F2 as the index
    """
    cell_start_as_numbers = convert_cell_ref_to_numbers_zero_based(
        workbook_range.CELL_START
    )
    cell_end_as_numbers = convert_cell_ref_to_numbers_one_based(workbook_range.CELL_END)

    range_data = get_range_data(sheet, cell_start_as_numbers, cell_end_as_numbers)

    range_values = range_data[1:, 1:]
    range_columns = range_data[0, 1:]
    range_index = range_data[1:, 0]

    return (
        pd.DataFrame(range_values, columns=range_columns, index=range_index)
        .dropna(axis=1, how="all")
        .replace("", np.nan)
        .dropna(axis=0, how="all")
    )


def load_disconnected_dataframe_from_workbook(
    index_workbook_range: ExcelWorkbookRange,
    column_workbook_range: ExcelWorkbookRange,
    data_workbook_range: ExcelWorkbookRange,
    sheet: Worksheet,
):
    """
    This function will read a workbook and return a dataframe.
    It will use different ranges for the index, columns and data of the datframe.
    This means you can set different rows as the index for the dataframe and different column names.
    This function will also convert the index / column range to a multi-index if:
    - the index spans more than 1 column
    - the columns span more more than 1 row.
    """
    (
        index_cell_start_as_numbers,
        index_cell_end_as_numbers,
    ) = get_cell_start_and_end_as_numbers(index_workbook_range)
    (
        column_cell_start_as_numbers,
        column_cell_end_as_numbers,
    ) = get_cell_start_and_end_as_numbers(column_workbook_range)
    (
        data_cell_start_as_numbers,
        data_cell_end_as_numbers,
    ) = get_cell_start_and_end_as_numbers(data_workbook_range)

    index_range = get_range_data(
        sheet, index_cell_start_as_numbers, index_cell_end_as_numbers
    )
    dataframe_index = process_index_range(index_range)

    column_range = get_range_data(
        sheet, column_cell_start_as_numbers, column_cell_end_as_numbers
    )
    dataframe_columns = process_index_range(column_range.T)
    data_range = get_range_data(
        sheet, data_cell_start_as_numbers, data_cell_end_as_numbers
    )

    assert data_range.shape[0] == len(
        dataframe_index
    ), "The data does not match the number of indices."

    assert data_range.shape[1] == len(
        dataframe_columns
    ), "The data does not match the number of columns."

    return pd.DataFrame(data_range, columns=dataframe_columns, index=dataframe_index)


def process_index_range(index_range: np.ndarray) -> Union[pd.Index, pd.MultiIndex]:
    index_is_multi_index = len(index_range[0]) > 1

    if index_is_multi_index:
        index_level_one = index_range[:, 0]
        index_level_two = index_range[:, 1]
        return create_multi_index_from_range(index_level_one, index_level_two)
    else:
        return pd.Index(index_range.flatten())


def create_multi_index_from_range(
    index_level_one: np.ndarray, index_level_two: np.ndarray
) -> pd.MultiIndex:
    index_level_one = np.unique(index_level_one[index_level_one != None])
    index_level_two = np.unique(index_level_two[index_level_two != None])

    return pd.MultiIndex.from_product([index_level_one, index_level_two])


def get_range_data(
    sheet: Worksheet,
    cell_start_as_numbers: CellReferenceNumbers,
    cell_end_as_numbers: CellReferenceNumbers,
) -> np.ndarray:
    row_start = cell_start_as_numbers.row
    column_start = cell_start_as_numbers.column
    row_end = cell_end_as_numbers.row
    column_end = cell_end_as_numbers.column

    return np.array(list(sheet.values))[row_start:row_end, column_start:column_end]


def load_workbook_column(
    workbook_range: ExcelWorkbookRange, sheet: Worksheet
) -> pd.Series:
    """
    This function will read in a column from an excel worksheet
    from the cell start to the cell end
    """
    cell_start_as_numbers = convert_cell_ref_to_numbers_zero_based(
        workbook_range.CELL_START
    )
    cell_end_as_numbers = convert_cell_ref_to_numbers_one_based(workbook_range.CELL_END)
    row_start = cell_start_as_numbers.row
    column_start = cell_start_as_numbers.column
    row_end = cell_end_as_numbers.row

    range_data = np.array(list(sheet.values))[
        row_start:row_end, column_start : column_start + 1
    ]

    range_values = range_data[1:]

    range_columns = range_data[0]

    return pd.DataFrame(range_values, columns=range_columns).squeeze()


def load_workbook_cell(cell_ref: str, sheet: Worksheet):
    cell_ref_as_numbers = convert_cell_ref_to_numbers_zero_based(cell_ref)

    return np.array(list(sheet.values))[
        cell_ref_as_numbers.row, cell_ref_as_numbers.column
    ]


def get_cell_start_and_end_as_numbers(
    workbook_range: ExcelWorkbookRange,
) -> tuple[CellReferenceNumbers, CellReferenceNumbers]:
    cell_start_as_numbers = convert_cell_ref_to_numbers_zero_based(
        workbook_range.CELL_START
    )
    cell_end_as_numbers = convert_cell_ref_to_numbers_one_based(workbook_range.CELL_END)

    return cell_start_as_numbers, cell_end_as_numbers
