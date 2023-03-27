import pandas as pd
from typing import cast
from asc_overview import params
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from asc_overview.utilities.convert_cell_ref_to_numbers import (
    convert_cell_ref_to_numbers_one_based,
)


def get_worksheet_from_workbook(workbook: Workbook, sheet_name: str) -> Worksheet:
    return cast(Worksheet, workbook[sheet_name])


def output_time_series_column_to_workbook(
    worksheet_to_output_to: Worksheet,
    new_time_series_column: pd.Series,
    cell_to_write_to: str,
    time_series_year: str = None,
) -> None:
    if time_series_year is None:
        time_series_year = params.PUBLICATION_YEAR

    cell_to_write_to_as_numbers = convert_cell_ref_to_numbers_one_based(
        cell_to_write_to
    )
    start_row = cell_to_write_to_as_numbers.row
    column_to_write_to = cell_to_write_to_as_numbers.column
    worksheet_to_output_to.cell(
        row=start_row, column=column_to_write_to, value=time_series_year
    )

    row_to_write_to = start_row + 1

    write_series_to_excel(
        new_time_series_column,
        worksheet_to_output_to,
        row_to_write_to,
        column_to_write_to,
    )


def write_series_to_excel(
    new_time_series_column: pd.Series,
    table_sheet: Worksheet,
    section_start_row: int,
    section_start_column: int,
):
    time_series_by_measure_rows = new_time_series_column.values.tolist()
    for row_index, cell_value in enumerate(
        time_series_by_measure_rows, section_start_row
    ):
        table_sheet.cell(row=row_index, column=section_start_column, value=cell_value)


def write_dataframe_to_excel(
    worksheet_to_output_to: Worksheet,
    df_to_output: pd.DataFrame,
    cell_to_write_to: str,
):
    cell_to_write_to_as_numbers = convert_cell_ref_to_numbers_one_based(
        cell_to_write_to
    )

    df_to_output_rows = df_to_output.values.tolist()
    for row_index, row_of_table in enumerate(
        df_to_output_rows, cell_to_write_to_as_numbers.row
    ):
        for column_index, cell_value in enumerate(
            row_of_table, cell_to_write_to_as_numbers.column
        ):
            worksheet_to_output_to.cell(
                row=row_index, column=column_index, value=cell_value
            )


def write_dataframe_to_excel_with_column_names(
    worksheet_to_output_to: Worksheet,
    df_to_output: pd.DataFrame,
    cell_to_write_to: str,
):
    cell_to_write_to_as_numbers = convert_cell_ref_to_numbers_one_based(
        cell_to_write_to
    )

    df_to_output_rows = df_to_output.values.tolist()
    for row_index, row_of_table in enumerate(
        df_to_output_rows, cell_to_write_to_as_numbers.row
    ):
        for column_index, cell_value in enumerate(
            row_of_table, cell_to_write_to_as_numbers.column
        ):
            worksheet_to_output_to.cell(
                row=row_index, column=column_index, value=cell_value
            )
