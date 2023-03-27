from asc_overview import params
from asc_overview import OUTPUT_TIME_SERIES_FILENAME
from asc_overview.input_data.download_publication import PATH_TO_DOWNLOADED_PUBLICATIONS
from openpyxl import load_workbook
from pathlib import Path


def load_ascs_publication():
    return load_workbook(
        str(Path(PATH_TO_DOWNLOADED_PUBLICATIONS) / params.PUBLICATIONS.ASCS.FILENAME)
    )


def load_ascfr_publication():
    return load_workbook(
        str(Path(PATH_TO_DOWNLOADED_PUBLICATIONS) / params.PUBLICATIONS.ASCFR.FILENAME)
    )


def load_dols_data():
    return load_workbook(
        str(
            Path(PATH_TO_DOWNLOADED_PUBLICATIONS)
            / params.PUBLICATIONS.DOLS_APPLICATIONS.FILENAME
        )
    )


def load_safeguarding_publication():
    return load_workbook(
        str(
            Path(PATH_TO_DOWNLOADED_PUBLICATIONS)
            / params.PUBLICATIONS.SAFEGUARDING.FILENAME
        )
    )


def load_overview_workbook():
    return load_workbook(
        str(Path(params.PATH_TO_OUTPUTS) / OUTPUT_TIME_SERIES_FILENAME)
    )


def load_ascof_workbook():
    return load_workbook(
        str(Path(PATH_TO_DOWNLOADED_PUBLICATIONS) / params.PUBLICATIONS.ASCOF.FILENAME)
    )


def load_workforce_publication():
    return load_workbook(
        str(
            Path(PATH_TO_DOWNLOADED_PUBLICATIONS)
            / params.PUBLICATIONS.WORKFORCE.FILENAME
        )
    )
