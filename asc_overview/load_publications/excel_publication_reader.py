import pandas as pd
from typing import Any
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from asc_overview.utilities.excel_config import (
    ExcelCellConfig,
    ExcelDisconnectedTableConfig,
    ExcelTableConfig,
)
from asc_overview.utilities.output_to_excel import get_worksheet_from_workbook
from asc_overview.utilities.load_from_excel import (
    load_disconnected_dataframe_from_workbook,
    load_workbook_cell,
    load_workbook_range_as_dataframe,
)


class ExcelPublicationReader:
    def __init__(self, wb: Workbook) -> None:
        self._wb: Workbook = wb

    def get_workbook(self) -> Workbook:
        return self._wb

    def get_sheet_from_worbook(self, sheet_name: str) -> Worksheet:
        return get_worksheet_from_workbook(self.get_workbook(), sheet_name)

    def get_cell_from_workbook(self, cell_config: ExcelCellConfig) -> Any:
        sheet = self.get_sheet_from_worbook(cell_config.SHEET_NAME)
        return load_workbook_cell(cell_config.CELL_REF, sheet)

    def get_range_from_workbook_as_dataframe(
        self, table_config: ExcelTableConfig
    ) -> pd.DataFrame:
        sheet = self.get_sheet_from_worbook(table_config.SHEET_NAME)
        return load_workbook_range_as_dataframe(table_config.WORKBOOK_RANGE, sheet)

    def get_disconnected_dataframe_from_workbook(
        self, table_config: ExcelDisconnectedTableConfig
    ) -> pd.DataFrame:
        sheet = self.get_sheet_from_worbook(table_config.SHEET_NAME)
        return load_disconnected_dataframe_from_workbook(
            table_config.INDEX_WORKBOOK_RANGE,
            table_config.COLUMN_WORKBOOK_RANGE,
            table_config.DATA_WORKBOOK_RANGE,
            sheet,
        )
