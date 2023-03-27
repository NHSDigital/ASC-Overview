from re import L
import numpy as np
import pandas as pd
import pandas.testing as pd_testing
import pytest
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from asc_overview.utilities.load_from_excel import (
    create_multi_index_from_range,
    load_disconnected_dataframe_from_workbook,
    load_workbook_cell,
    load_workbook_column,
    load_workbook_range_as_dataframe,
)
from asc_overview.utilities.excel_config import ExcelWorkbookRange


@pytest.fixture
def test_workbook() -> Workbook:
    return load_workbook("./tests/test_data/test.xlsx")


@pytest.fixture
def test_sheet(test_workbook: Workbook) -> Worksheet:
    return test_workbook["Sheet1"]  # type: ignore


def test_load_workbook_range_as_dataframe(test_sheet: Worksheet):
    expected_df = pd.DataFrame(
        {
            "column_1": ["foo", "boo"],
            "column_2": ["bar", "far"],
            "column_3": ["other", "value"],
        },
        index=["index_1", "index_2"],
    )

    workbook_range_dict = {"CELL_END": "E5", "CELL_START": "B3"}
    workbook_range = ExcelWorkbookRange(workbook_range_dict)
    actual_df = load_workbook_range_as_dataframe(workbook_range, test_sheet)

    pd_testing.assert_frame_equal(expected_df, actual_df)


def test_load_workbook_range_as_dataframe__with_cell_end(test_sheet: Worksheet):
    expected_df = pd.DataFrame(
        {
            "column_1": ["foo", "boo"],
            "column_2": ["bar", "far"],
        },
        index=["index_1", "index_2"],
    )

    workbook_range_dict = {"CELL_END": "D5", "CELL_START": "B3"}
    workbook_range = ExcelWorkbookRange(workbook_range_dict)
    actual_df = load_workbook_range_as_dataframe(workbook_range, test_sheet)

    pd_testing.assert_frame_equal(expected_df, actual_df)


def test_load_disconnected_dataframe_from_workbook(test_sheet: Worksheet):
    expected_df = pd.DataFrame(
        {
            "disconnected_column_1": ["foo", "boo"],
            "disconnected_column_2": ["bar", "far"],
            "disconnected_column_3": ["other", "value"],
        },
        index=["disconnected_index_1", "disconnected_index_2"],
    )

    index_workbook_range_dict = {"CELL_START": "A7", "CELL_END": "A8"}
    index_workbook_range = ExcelWorkbookRange(index_workbook_range_dict)

    column_workbook_range_dict = {"CELL_START": "C1", "CELL_END": "E1"}
    column_workbook_range = ExcelWorkbookRange(column_workbook_range_dict)

    data_workbook_range_dict = {"CELL_START": "C4", "CELL_END": "E5"}
    data_workbook_range = ExcelWorkbookRange(data_workbook_range_dict)

    actual_df = load_disconnected_dataframe_from_workbook(
        index_workbook_range, column_workbook_range, data_workbook_range, test_sheet
    )

    pd_testing.assert_frame_equal(actual_df, expected_df)


def test_load_disconnected_dataframe_from_workbook__throws_error_for_index(
    test_sheet: Worksheet,
):
    index_workbook_range_dict = {"CELL_START": "A7", "CELL_END": "A9"}
    index_workbook_range = ExcelWorkbookRange(index_workbook_range_dict)

    column_workbook_range_dict = {"CELL_START": "C1", "CELL_END": "E1"}
    column_workbook_range = ExcelWorkbookRange(column_workbook_range_dict)

    data_workbook_range_dict = {"CELL_START": "C4", "CELL_END": "E5"}
    data_workbook_range = ExcelWorkbookRange(data_workbook_range_dict)

    with pytest.raises(AssertionError) as err:
        load_disconnected_dataframe_from_workbook(
            index_workbook_range, column_workbook_range, data_workbook_range, test_sheet
        )

    assert "indices" in str(err.value)


def test_load_disconnected_dataframe_from_workbook__throws_error_for_column(
    test_sheet: Worksheet,
):
    index_workbook_range_dict = {"CELL_START": "A7", "CELL_END": "A8"}
    index_workbook_range = ExcelWorkbookRange(index_workbook_range_dict)

    column_workbook_range_dict = {"CELL_START": "C1", "CELL_END": "F1"}
    column_workbook_range = ExcelWorkbookRange(column_workbook_range_dict)

    data_workbook_range_dict = {"CELL_START": "C4", "CELL_END": "E5"}
    data_workbook_range = ExcelWorkbookRange(data_workbook_range_dict)

    with pytest.raises(AssertionError) as err:
        load_disconnected_dataframe_from_workbook(
            index_workbook_range, column_workbook_range, data_workbook_range, test_sheet
        )

    assert "columns" in str(err.value)


def test_load_disconnected_dataframe_from_workbook_multiindex_index(
    test_sheet: Worksheet,
):
    index_workbook_range_dict = {"CELL_START": "H4", "CELL_END": "I7"}
    index_workbook_range = ExcelWorkbookRange(index_workbook_range_dict)

    column_workbook_range_dict = {"CELL_START": "J3", "CELL_END": "L3"}
    column_workbook_range = ExcelWorkbookRange(column_workbook_range_dict)

    data_workbook_range_dict = {"CELL_START": "J4", "CELL_END": "L7"}
    data_workbook_range = ExcelWorkbookRange(data_workbook_range_dict)

    expected_df = pd.DataFrame(
        {
            "column_1": ["value_1", "value_4", "value_7", "value_10"],
            "column_2": ["value_2", "value_5", "value_8", "value_11"],
            "column_3": ["value_3", "value_6", "value_9", "value_12"],
        },
        index=pd.MultiIndex.from_product(
            [["multi_index_1", "multi_index_2"], ["lower_level_1", "lower_level_2"]]
        ),
    )

    actual_df = load_disconnected_dataframe_from_workbook(
        index_workbook_range, column_workbook_range, data_workbook_range, test_sheet
    )

    pd_testing.assert_frame_equal(actual_df, expected_df)


def test_load_disconnected_dataframe_from_workbook_multiindex_column(
    test_sheet: Worksheet,
):
    index_workbook_range_dict = {"CELL_START": "B14", "CELL_END": "B14"}
    index_workbook_range = ExcelWorkbookRange(index_workbook_range_dict)

    column_workbook_range_dict = {"CELL_START": "C12", "CELL_END": "F13"}
    column_workbook_range = ExcelWorkbookRange(column_workbook_range_dict)

    data_workbook_range_dict = {"CELL_START": "C14", "CELL_END": "F14"}
    data_workbook_range = ExcelWorkbookRange(data_workbook_range_dict)

    expected_df = pd.DataFrame(
        [["value_1", "value_2", "value_3", "value_4"]],
        index=["index_1"],
        columns=pd.MultiIndex.from_product(
            [
                ["multi_index_column_1", "multi_index_column_2"],
                ["low_level_1", "low_level_2"],
            ]
        ),
    )

    actual_df = load_disconnected_dataframe_from_workbook(
        index_workbook_range, column_workbook_range, data_workbook_range, test_sheet
    )

    pd_testing.assert_frame_equal(actual_df, expected_df)


def test_create_multi_index_from_range():
    index_level_one = np.array(["multi_index_1", "multi_index_2"])
    index_level_two = np.array(["lower_level_1", "lower_level_2", "lower_level_3"])

    expected_index = pd.MultiIndex.from_product(
        [
            [
                "multi_index_1",
                "multi_index_2",
            ],
            ["lower_level_1", "lower_level_2", "lower_level_3"],
        ]
    )

    actual_index = create_multi_index_from_range(index_level_one, index_level_two)

    pd_testing.assert_index_equal(actual_index, expected_index)


def test_load_workbook_column(test_sheet: Worksheet):
    expected_series = pd.Series(["bar", "far"], name="column_2")

    workbook_range_dict = {"CELL_START": "D3", "CELL_END": "D5"}
    workbook_range = ExcelWorkbookRange(workbook_range_dict)

    actual_series = load_workbook_column(workbook_range, test_sheet)

    pd_testing.assert_series_equal(actual_series, expected_series)


def test_load_workbook_cell(test_sheet: Worksheet):
    expected_value = "foo"

    actual_value = load_workbook_cell("C4", test_sheet)

    assert actual_value == expected_value
