import pytest
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from asc_overview import params
from asc_overview.load_publications.publications import Publications
from asc_overview.time_series_tables.new_requests_table.new_requests_table import (
    NewRequestsForSupport,
    calculate_dols_applications_received,
    calculate_new_requests_percent_growth_by_age,
    calculate_safeguarding_concerns_raised,
    create_new_time_series_column,
    calculate_new_requests_for_support_section,
    reindex_rows,
)
from asc_overview.utilities.load_from_excel import load_workbook_range_as_dataframe


@pytest.fixture
def existing_time_series_data():
    wb = load_workbook(f"./{params.ASC_OVERVIEW_TEMPLATE_FILE_NAME}")
    table_1_sheet = wb["Table 1"]
    return load_workbook_range_as_dataframe(
        params.NEW_REQUESTS_TABLE.OVERVIEW.WORKBOOK_RANGE, table_1_sheet
    )


@pytest.fixture
def new_requests_by_age(publications: Publications):
    return publications.ascfr.get_new_requests_by_age()


@pytest.fixture
def dols_applications_received(publications: Publications):
    return publications.dols.get_applications_received()


@pytest.fixture
def new_time_series_column(existing_time_series_data, publications):
    return create_new_time_series_column(existing_time_series_data, publications)


def test_create_new_time_series_column(existing_time_series_data, publications):
    expected_series = pd.Series(
        data=[
            577765.0,
            1337875.0,
            0.154,
            0.021,
            33992830.0,
            10464020.0,
            1710.0,
            12815.0,
            498260.0,
            256610.0,
            0.048,
            -0.028,
        ],
        index=params.NEW_REQUESTS_TABLE.get_row_order(),
        name=params.PUBLICATION_YEAR,
    )

    actual_series = create_new_time_series_column(
        existing_time_series_data, publications
    )

    pd.testing.assert_series_equal(expected_series, actual_series)


def test_calculate_dols_applications_received(
    publications: Publications, existing_time_series_data: pd.DataFrame
):
    expected_series = pd.Series(
        [256610, -0.02777146321133591],
        index=[
            params.NEW_REQUESTS_TABLE.ROW_NAMES.DOLS_APPLICATIONS_RECEIVED_NAME,
            params.NEW_REQUESTS_TABLE.ROW_NAMES.DOLS_YEAR_ON_YEAR_APPLICATIONS_RECEIVED_NAME,
        ],
        name=params.PUBLICATION_YEAR,
    )

    actual_series = calculate_dols_applications_received(
        publications, existing_time_series_data
    )

    pd.testing.assert_series_equal(expected_series, actual_series)


def test_calculate_safeguarding_concerns_raised(
    publications: Publications, existing_time_series_data: pd.DataFrame
):
    expected_series = pd.Series(
        [498260, 0.04773319875515182],
        index=[
            params.NEW_REQUESTS_TABLE.ROW_NAMES.SAFEGUARDING_CONCERNS_RAISED_NAME,
            params.NEW_REQUESTS_TABLE.ROW_NAMES.SAFEGUARDING_YEAR_ON_YEAR_CONCERNS_RAISED_NAME,
        ],
        name=params.PUBLICATION_YEAR,
    )

    actual_series = calculate_safeguarding_concerns_raised(
        publications, existing_time_series_data
    )

    pd.testing.assert_series_equal(expected_series, actual_series)


def test_calculate_new_requests_for_support_section(
    existing_time_series_data: pd.DataFrame, publications: Publications
):
    new_requests_for_support_by_age = pd.Series(
        data=[
            1710.0,
            577765.0,
            33992830.0,
            12815.0,
            1337875.0,
            10464020.0,
        ],
        index=params.NEW_REQUESTS_TABLE.ASCFR_NEW_INDEX,
        name=params.PUBLICATION_YEAR,
    )
    new_requests_for_support_proportion_growth_by_age = pd.Series(
        data=[
            0.153984,
            0.021232,
        ],
        name=params.PUBLICATION_YEAR,
        index=[
            params.NEW_REQUESTS_TABLE.ROW_NAMES.ASCFR_GROWTH_18_64_SINCE_2015_16_NAME,
            params.NEW_REQUESTS_TABLE.ROW_NAMES.ASCFR_GROWTH_65_PLUS_SINCE_2015_16_NAME,
        ],
    )

    expected_result = NewRequestsForSupport(
        new_requests_for_support_by_age,
        new_requests_for_support_proportion_growth_by_age,
    )

    actual_result = calculate_new_requests_for_support_section(
        existing_time_series_data, publications
    )

    pd.testing.assert_series_equal(
        expected_result.new_requests_for_support_by_age,
        actual_result.new_requests_for_support_by_age,
    )
    pd.testing.assert_series_equal(
        expected_result.new_requests_for_support_proportion_growth_by_age,
        actual_result.new_requests_for_support_proportion_growth_by_age,
    )


def test_calculate_new_requests_percent_growth_by_age(
    existing_time_series_data, new_requests_by_age
):
    expected_series = pd.Series(
        data=[
            0.153984,
            0.021232,
        ],
        name=params.PUBLICATION_YEAR,
        index=[
            params.NEW_REQUESTS_TABLE.ROW_NAMES.ASCFR_GROWTH_18_64_SINCE_2015_16_NAME,
            params.NEW_REQUESTS_TABLE.ROW_NAMES.ASCFR_GROWTH_65_PLUS_SINCE_2015_16_NAME,
        ],
    )

    actual_series = calculate_new_requests_percent_growth_by_age(
        new_requests_by_age, existing_time_series_data
    )

    pd.testing.assert_series_equal(expected_series, actual_series)
