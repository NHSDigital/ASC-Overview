import pandas as pd
import pandas.testing as pd_testing
import pytest
from _pytest.monkeypatch import MonkeyPatch
from pytest_mock import MockerFixture
from openpyxl import load_workbook, Workbook
from asc_overview.load_publications.ascof_reader.ascof_loading_config import (
    OutcomesByLa,
)
from asc_overview.load_publications.ascof_reader.ascof_reader import AscofReader
from asc_overview.load_publications.load_publications import load_ascof_workbook
from asc_overview import params


@pytest.fixture
def ascof_reader() -> AscofReader:
    return AscofReader(load_ascof_workbook())


@pytest.fixture(scope="module")
def workbook_with_carers_data() -> Workbook:
    return load_workbook("./tests/test_data/ascof_18_19_with_carers.xlsx")


@pytest.fixture(scope="module")
def workbook_outcomes_by_la() -> Workbook:
    return load_workbook("./tests/test_data/outcomes_by_la_test_data.xlsx")


def test_create_age_series_from_config(ascof_reader: AscofReader):
    expected_series = pd.Series(
        {"18 to 64": 69.9, "65 and over": 66.1}, name=params.PUBLICATION_YEAR
    )

    actual_series = ascof_reader.create_age_series_from_config(
        params.ASCOF_LOADING_CONFIG.SERVICE_USER_SATISFACTION
    )

    pd_testing.assert_series_equal(actual_series, expected_series)


def test_get_service_user_percentage_satisfaction_by_age(ascof_reader: AscofReader):
    expected_series = pd.Series(
        [0.699, 0.661],
        index=["18 to 64", "65 and over"],
        name=params.PUBLICATION_YEAR,
    )

    actual_series = ascof_reader.get_service_user_percentage_satisfaction_by_age()

    pd_testing.assert_series_equal(actual_series, expected_series)


def test_get_service_user_quality_of_life_score_by_age(ascof_reader: AscofReader):
    expected_series = pd.Series(
        [19.2, 18.8], index=["18 to 64", "65 and over"], name=params.PUBLICATION_YEAR
    )

    actual_series = ascof_reader.get_service_user_quality_of_life_score_by_age()

    pd_testing.assert_series_equal(actual_series, expected_series)


def test_get_carer_percentage_satisfaction_by_age__given_no_carer_data(
    ascof_reader: AscofReader,
):
    expected_series = pd.Series(
        ["N/A", "N/A"],
        index=[
            params.SOCIAL_CARE_EXPERIENCE_TABLE.ROW_NAMES.CARER_SATISFACTION_18_64,
            params.SOCIAL_CARE_EXPERIENCE_TABLE.ROW_NAMES.CARER_SATISFACTION_65_AND_OVER,
        ],
        name=params.PUBLICATION_YEAR,
    )

    actual_series = ascof_reader.get_carer_percentage_satisfaction_by_age()

    pd_testing.assert_series_equal(actual_series, expected_series)


def test_get_carer_percentage_satisfaction_by_age__given_no_carer_data_throws_error(
    ascof_reader: AscofReader, mocker: MockerFixture
):
    mocker.patch(
        "asc_overview.load_publications.ascof_reader.ascof_reader.params.LOAD_CARERS_DATA",
        return_value=True,
    )
    with pytest.raises(AssertionError) as err:
        ascof_reader.get_carer_percentage_satisfaction_by_age()

    assert "No carers data found!" in str(err.value)


def test_get_carer_percentage_satisfaction_by_age__given_carer_data(
    ascof_reader: AscofReader,
    mocker: MockerFixture,
    workbook_with_carers_data: Workbook,
):
    mocker.patch(
        "asc_overview.load_publications.ascof_reader.ascof_reader.params.LOAD_CARERS_DATA",
        return_value=True,
    )
    mocker.patch(
        "asc_overview.load_publications.ascof_reader.ascof_reader.AscofReader.get_workbook",
        return_value=workbook_with_carers_data,
    )

    expected_series = pd.Series(
        [0.361, 0.410],
        index=[
            "18 to 64",
            "65 and over",
        ],
        name=params.PUBLICATION_YEAR,
    )

    actual_series = ascof_reader.get_carer_percentage_satisfaction_by_age()

    pd_testing.assert_series_equal(actual_series, expected_series)


def test_get_carer_quality_of_life_score_by_age__given_no_carer_data(
    ascof_reader: AscofReader,
):
    expected_series = pd.Series(
        ["N/A", "N/A"],
        index=["18 to 64", "65 and over"],
        name=params.PUBLICATION_YEAR,
    )

    actual_series = ascof_reader.get_carer_quality_of_life_score_by_age()

    pd_testing.assert_series_equal(actual_series, expected_series)


def test_get_carer_quality_of_life_score_by_age__given_carer_data(
    ascof_reader: AscofReader,
    mocker: MockerFixture,
    workbook_with_carers_data: Workbook,
):
    mocker.patch(
        "asc_overview.load_publications.ascof_reader.ascof_reader.params.LOAD_CARERS_DATA",
        return_value=True,
    )
    mocker.patch(
        "asc_overview.load_publications.ascof_reader.ascof_reader.AscofReader.get_workbook",
        return_value=workbook_with_carers_data,
    )

    expected_series = pd.Series(
        [7.2, 7.8], index=["18 to 64", "65 and over"], name=params.PUBLICATION_YEAR
    )

    actual_series = ascof_reader.get_carer_quality_of_life_score_by_age()

    pd_testing.assert_series_equal(actual_series, expected_series)


def test_get_carer_quality_of_life_by_age__given_no_carer_data_throws_error(
    ascof_reader: AscofReader, mocker: MockerFixture
):
    mocker.patch(
        "asc_overview.load_publications.ascof_reader.ascof_reader.params.LOAD_CARERS_DATA",
        return_value=True,
    )
    with pytest.raises(AssertionError) as err:
        ascof_reader.get_carer_quality_of_life_score_by_age()

    assert "No carers data found!" in str(err.value)


def test_get_carer_social_contact_by_age__given_no_carer_data(
    ascof_reader: AscofReader,
):
    expected_series = pd.Series(
        ["N/A", "N/A"],
        index=["18 to 64", "65 and over"],
        name=params.PUBLICATION_YEAR,
    )

    actual_series = ascof_reader.get_carer_social_contact_by_age()

    pd_testing.assert_series_equal(actual_series, expected_series)


def test_get_carer_social_contact_by_age__given_no_carer_data_throws_error(
    ascof_reader: AscofReader, mocker: MockerFixture
):
    mocker.patch(
        "asc_overview.load_publications.ascof_reader.ascof_reader.params.LOAD_CARERS_DATA",
        return_value=True,
    )
    with pytest.raises(AssertionError) as err:
        ascof_reader.get_carer_social_contact_by_age()

    assert "No carers data found!" in str(err.value)


def test_get_carer_social_contact_by_age__given_carer_data(
    ascof_reader: AscofReader,
    mocker: MockerFixture,
    workbook_with_carers_data: Workbook,
):
    mocker.patch(
        "asc_overview.load_publications.ascof_reader.ascof_reader.params.LOAD_CARERS_DATA",
        return_value=True,
    )
    mocker.patch(
        "asc_overview.load_publications.ascof_reader.ascof_reader.AscofReader.get_workbook",
        return_value=workbook_with_carers_data,
    )

    expected_series = pd.Series(
        [0.295, 0.345],
        index=["18 to 64", "65 and over"],
        name=params.PUBLICATION_YEAR,
    )

    actual_series = ascof_reader.get_carer_social_contact_by_age()

    pd_testing.assert_series_equal(actual_series, expected_series)


def test_get_service_user_safety(ascof_reader: AscofReader):
    expected_series = pd.Series(
        [0.709, 0.756], index=["18 to 64", "65 and over"], name=params.PUBLICATION_YEAR
    )

    actual_series = ascof_reader.get_service_user_safety()

    pd_testing.assert_series_equal(actual_series, expected_series)


def test_get_service_user_social_contact_by_age(ascof_reader: AscofReader):
    expected_series = pd.Series(
        [0.368, 0.327], index=["18 to 64", "65 and over"], name=params.PUBLICATION_YEAR
    )

    actual_series = ascof_reader.get_service_user_social_contact_by_age()

    pd_testing.assert_series_equal(actual_series, expected_series)


def test_get_service_user_social_contact_by_la(
    ascof_reader: AscofReader,
    workbook_outcomes_by_la: Workbook,
    mocker: MockerFixture,
    monkeypatch: MonkeyPatch,
):
    mocker.patch(
        "asc_overview.load_publications.ascof_reader.ascof_reader.AscofReader.get_workbook",
        return_value=workbook_outcomes_by_la,
    )
    social_contact_by_la_dict = {
        "SHEET_NAME": "1I(1)",
        "RANGE_INDEX": {"CELL_START": "B10", "CELL_END": "B13"},
        "RANGE_18_64": {"CELL_START": "O10", "CELL_END": "B13"},
        "RANGE_65_AND_OVER": {"CELL_START": "R10", "CELL_END": "B13"},
    }
    monkeypatch.setattr(
        params.ASCOF_LOADING_CONFIG,
        "SOCIAL_CONTACT_BY_LA",
        OutcomesByLa(social_contact_by_la_dict),
    )

    df_expected = pd.DataFrame(
        {"18 to 64": [54.3, 42.8, "[x]"], "65 and over": [45.9, 41.9, "[x]"]},
        index=pd.Index(["la_one", "la_two", "la_three"], name="la_index"),
    )

    df_actual = ascof_reader.get_service_user_socal_contact_by_la()

    pd_testing.assert_frame_equal(df_actual, df_expected)


def test_get_service_user_feel_safe_by_la(
    ascof_reader: AscofReader,
    workbook_outcomes_by_la: Workbook,
    mocker: MockerFixture,
    monkeypatch: MonkeyPatch,
):
    mocker.patch(
        "asc_overview.load_publications.ascof_reader.ascof_reader.AscofReader.get_workbook",
        return_value=workbook_outcomes_by_la,
    )
    feel_safe_by_la_dict = {
        "SHEET_NAME": "4A",
        "RANGE_INDEX": {"CELL_START": "B6", "CELL_END": "B9"},
        "RANGE_18_64": {"CELL_START": "O6", "CELL_END": "O9"},
        "RANGE_65_AND_OVER": {"CELL_START": "R6", "CELL_END": "R9"},
    }
    monkeypatch.setattr(
        params.ASCOF_LOADING_CONFIG,
        "FEEL_SAFE_BY_LA",
        OutcomesByLa(feel_safe_by_la_dict),
    )

    df_expected = pd.DataFrame(
        {"18 to 64": [1.0, 2.0, "[x]"], "65 and over": [3.0, 4.0, "[x]"]},
        index=pd.Index(["la_one", "la_two", "la_three"], name="la_index"),
    )

    df_actual = ascof_reader.get_service_user_feel_safe_by_la()

    pd_testing.assert_frame_equal(df_actual, df_expected)
