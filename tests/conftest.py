import pytest
from openpyxl import load_workbook
from pytest_mock import MockFixture
from asc_overview import params, OUTPUT_TIME_SERIES_FILENAME
from asc_overview.create_publication import (
    output_all_cms_tables,
)
from asc_overview.load_publications.acfr_reader.ascfr_reader import AscfrReader
from asc_overview.load_publications.ascof_reader.ascof_reader import AscofReader
from asc_overview.load_publications.ascs_reader.ascs_reader import AscsReader
from asc_overview.load_publications.dols_reader.dols_reader import DolsReader
from asc_overview.load_publications.ons_reader import OnsReader
from asc_overview.load_publications.safeguarding_reader.safeguarding_reader import (
    SafeguardingReader,
)
from asc_overview.load_publications.overview_reader.overview_reader import (
    OverviewReader,
)
from asc_overview.load_publications.workforce_reader.workforce_reader import (
    WorkforceReader,
)
from asc_overview.load_publications.publications import Publications
from asc_overview.load_publications.load_publications import (
    load_ascfr_publication,
    load_ascof_workbook,
    load_ascs_publication,
    load_dols_data,
    load_safeguarding_publication,
    load_workforce_publication,
)
from asc_overview.time_series_tables.la_expenditure_table.la_expenditure_table import (
    add_new_la_expenditure_time_series_column_to_workbook,
)
from asc_overview.time_series_tables.new_requests_table.new_requests_table import (
    add_new_requests_time_series_column_to_workbook,
)
from asc_overview.time_series_tables.short_term_care.short_term_care_table import (
    add_new_short_term_care_time_series_column_to_workbook,
)
from asc_overview.time_series_tables.long_term_care_table.long_term_care_table import (
    add_new_long_term_care_time_series_column_to_workbook,
)
from asc_overview.time_series_tables.social_care_experience_table.social_care_experience_table import (
    add_new_social_care_experience_time_series_column_to_workbook,
)
from asc_overview.time_series_tables.outcomes_table.outcomes_table import (
    add_new_outcomes_time_series_column_to_workbook,
)
from asc_overview.time_series_tables.more_outcomes_table.more_outcomes_table import (
    add_new_more_outcomes_time_series_column_to_workbook,
)
from asc_overview.la_breakdown_tables.outcomes_by_la_table.outcomes_by_la_table import (
    add_outcomes_by_la_to_workbook,
)
from asc_overview.time_series_tables.workforce_table.workforce_table import (
    add_new_workforce_time_series_column_to_workbook,
)


@pytest.fixture(scope="session")
def ascfr_reader():
    return AscfrReader(load_ascfr_publication())


@pytest.fixture(scope="session")
def ascof_reader():
    return AscofReader(load_ascof_workbook())


@pytest.fixture(scope="session")
def ascs_reader():
    return AscsReader(load_ascs_publication())


@pytest.fixture(scope="session")
def dols_reader():
    return DolsReader(load_dols_data())


@pytest.fixture(scope="session")
def safeguarding_reader():
    return SafeguardingReader(load_safeguarding_publication())


@pytest.fixture
def overview_reader(save_to_excel: str):
    return OverviewReader(load_workbook(save_to_excel))


@pytest.fixture(scope="session")
def workforce_reader():
    return WorkforceReader(load_workforce_publication())


@pytest.fixture
def ons_reader(mocker: MockFixture):
    ons_reader_mock = mocker.patch(
        "asc_overview.load_publications.ons_reader.OnsReader"
    )
    ons_reader_instance_mock = ons_reader_mock.return_value
    ons_reader_instance_mock.get_population_data_18_64.return_value = 33992831
    ons_reader_instance_mock.get_population_data_65_plus.return_value = 10464019
    return ons_reader_instance_mock


@pytest.fixture
def publications(
    ascfr_reader: AscfrReader,
    ascof_reader: AscofReader,
    ascs_reader: AscsReader,
    dols_reader: DolsReader,
    safeguarding_reader: SafeguardingReader,
    ons_reader: OnsReader,
    workforce_reader: WorkforceReader,
):
    return Publications(
        ascfr=ascfr_reader,
        ascof=ascof_reader,
        ascs=ascs_reader,
        dols=dols_reader,
        safeguarding=safeguarding_reader,
        ons=ons_reader,
        workforce=workforce_reader,
    )


@pytest.fixture
def my_temp_path(tmp_path_factory) -> str:
    return tmp_path_factory.mktemp("temp")


@pytest.fixture
def save_to_excel(my_temp_path: str, publications: Publications) -> str:
    output_path = f"{my_temp_path}/{OUTPUT_TIME_SERIES_FILENAME}"
    template_workbook = load_workbook(f"./{params.ASC_OVERVIEW_TEMPLATE_FILE_NAME}")
    add_new_requests_time_series_column_to_workbook(template_workbook, publications)
    add_new_short_term_care_time_series_column_to_workbook(
        template_workbook, publications
    )
    add_new_long_term_care_time_series_column_to_workbook(
        template_workbook, publications
    )
    add_new_la_expenditure_time_series_column_to_workbook(
        template_workbook, publications
    )
    add_new_social_care_experience_time_series_column_to_workbook(
        template_workbook, publications
    )
    add_new_outcomes_time_series_column_to_workbook(template_workbook, publications)
    add_new_more_outcomes_time_series_column_to_workbook(
        template_workbook, publications
    )
    add_outcomes_by_la_to_workbook(template_workbook, publications)
    add_new_workforce_time_series_column_to_workbook(template_workbook, publications)
    template_workbook.save(output_path)

    return output_path


@pytest.fixture
def save_cms_to_excel(overview_reader: OverviewReader, my_temp_path: str) -> str:
    output_all_cms_tables(overview_reader, my_temp_path)

    return my_temp_path
